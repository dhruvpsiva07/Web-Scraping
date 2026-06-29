import re
import time
import openpyxl
import requests
from googlesearch import search
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import warnings
import logging

# Configure logging and warnings
warnings.filterwarnings("ignore")
logging.getLogger('WDM').setLevel(logging.NOTSET)
logging.getLogger('urllib3').setLevel(logging.WARNING)

def get_first_link(query):
    try:
        search_results = search(query, num_results=1, sleep_interval=5)
        return next(search_results, None)
    except Exception as e:
        print(f"\n Google search failed for {query}: {str(e)[:100]}...")
        return None

def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

def find_contact_page(url):
    driver = setup_driver()
    try:
        driver.get(url)
        time.sleep(2)
        
        # First try to find contact link in page content
        contact_keywords = ["contact", "directory"]
        links = driver.find_elements(By.TAG_NAME, "a")
        
        contact_links = []
        for link in links:
            try:
                href = link.get_attribute("href")
                if href and urlparse(href).netloc == urlparse(url).netloc:
                    text = link.text.lower()
                    if any(keyword in text for keyword in contact_keywords):
                        contact_links.append((href, text))
            except:
                continue
        
        # Prioritize links with "contact" in them
        if contact_links:
            contact_links.sort(key=lambda x: len(x[1]))  # Prefer shorter links
            return contact_links[0][0]
        
        # If no contact link found, try common contact page paths
        common_paths = ["/contact", "/contact-us", "/contact.aspx", "/contact.php"]
        base_url = urlparse(url).scheme + "://" + urlparse(url).netloc
        
        for path in common_paths:
            contact_url = base_url + path
            try:
                driver.get(contact_url)
                time.sleep(1)
                if "404" not in driver.title.lower() and "not found" not in driver.title.lower():
                    return contact_url
            except:
                continue
        
        return url
    except Exception as e:
        print(f"\n Contact page error for {url}: {str(e)[:100]}...")
        return url
    finally:
        driver.quit()

def extract_emails(soup, url):
    emails = set()
    
    # Look for mailto links first (most reliable)
    for link in soup.find_all('a', href=True):
        if 'mailto:' in link['href'].lower():
            email = link['href'].split('mailto:')[-1].split('?')[0].strip()
            if re.match(r"[^@]+@[^@]+\.[^@]+", email):
                emails.add(email)
    
    # Then look in text (less reliable)
    text_emails = set(re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", soup.get_text()))
    emails.update(text_emails)
    
    # Filter out common non-person emails
    filtered_emails = set()
    domain = urlparse(url).netloc
    for email in emails:
        if not any(word in email.lower() for word in ['info', 'contact', 'help', 'support', 'webmaster', 'admin']):
            if domain.split('.')[-2] in email:  # Only keep emails matching school domain
                filtered_emails.add(email)
    
    return filtered_emails

def extract_phones(soup):
    phones = set()
    # Common US phone number patterns
    patterns = [
        r"\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}",  # (123) 456-7890 or 123-456-7890
        r"\d{3}[-.\s]\d{3}[-.\s]\d{4}",
        r"\(\d{3}\)\s?\d{3}-\d{4}",
        r"\d{3}\.\d{3}\.\d{4}"
    ]
    
    # Look in specific elements first
    for element in soup.find_all(['a', 'span', 'div', 'p']):
        text = element.get_text()
        for pattern in patterns:
            found = re.findall(pattern, text)
            if found:
                phones.update(found)
    
    return phones

def extract_contact_name(soup):
    # Focus on department contacts rather than individuals
    departments = set()
    
    # Look for headings that might contain department names
    for heading in soup.find_all(['h2', 'h3', 'h4']):
        text = heading.get_text().strip()
        if any(word in text.lower() for word in ['office', 'department', 'admissions', 'registrar', 'financial']):
            departments.add(text)
    
    if departments:
        return ", ".join(sorted(departments))
    return "General Contact"

def extract_contact_info(url):
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")
        
        emails = extract_emails(soup, url)
        phones = extract_phones(soup)
        contact = extract_contact_name(soup)
        
        return {
            'email': ", ".join(emails) if emails else "Not found",
            'phone': ", ".join(phones) if phones else "Not found",
            'contact': contact,
            'url': url
        }
    except Exception as e:
        print(f"\n Extraction error for {url}: {str(e)[:100]}...")
        return {'email': 'Error', 'phone': 'Error', 'contact': 'Error', 'url': url}

def main():
    # Load input Excel file
    try:
        input_wb = openpyxl.load_workbook('Baltimore_School_Names_and_Queries.xlsx')  # Change to your input filename
        input_ws = input_wb.active
    except Exception as e:
        print(f"\n Error loading input Excel file: {e}")
        return

    # Prepare output workbook
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "Scraped Results"
    output_ws.append(["School Name", "Google Query", "Result URL", "Email", "Phone", "Contact/Department", "Status"])

    # Process each row in input file
    for row in input_ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        school_name, google_query = row[0], row[1]
        print(f"\n Processing: {school_name}")
        
        status = "Success"
        result_url = ""
        email = ""
        phone = ""
        contact = ""
        
        try:
            # Get first result from Google
            result_url = get_first_link(google_query)
            if not result_url:
                status = "No results found"
                output_ws.append([school_name, google_query, "", "", "", "", status])
                continue
            
            # Find contact page
            contact_page = find_contact_page(result_url)
            
            # Extract contact info
            contact_info = extract_contact_info(contact_page)
            
            email = contact_info['email']
            phone = contact_info['phone']
            contact = contact_info['contact']
            
        except Exception as e:
            status = f"Error: {str(e)[:100]}..."
        
        # Write results to output
        output_ws.append([
            school_name,
            google_query,
            result_url,
            email,
            phone,
            contact,
            status
        ])
        
        # Save progress after each school
        output_wb.save("school_contacts_results.xlsx")
        time.sleep(5)  # Be polite with delays
    
    print("\n Scraping completed! Results saved to school_contacts_results.xlsx")

if __name__ == "__main__":
    main()